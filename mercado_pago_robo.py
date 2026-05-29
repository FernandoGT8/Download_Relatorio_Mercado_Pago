from __future__ import annotations

import asyncio
import hashlib
import os
import re
import sys
import time
import unicodedata
from copy import copy
from dataclasses import dataclass
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from playwright.async_api import (
    BrowserContext,
    ElementHandle,
    Locator,
    Page,
    TimeoutError as PlaywrightTimeoutError,
    async_playwright,
)
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.traceback import install


# =============================================================================
# CONFIGURAÇÕES
# =============================================================================

install(show_locals=True)
console = Console(log_path=False)


def env_bool(nome: str, padrao: bool = False) -> bool:
    valor = os.getenv(nome)
    if valor is None:
        return padrao
    return valor.strip().lower() in {"1", "true", "t", "sim", "s", "yes", "y"}


def env_int(nome: str, padrao: int, minimo: int | None = None) -> int:
    valor = os.getenv(nome)
    if valor is None:
        return padrao

    try:
        numero = int(valor)
    except ValueError:
        return padrao

    if minimo is not None:
        numero = max(minimo, numero)
    return numero


MODO_DETALHADO = env_bool("MODO_DETALHADO", False)
MAX_TENTATIVAS_ITEM = env_int("MAX_TENTATIVAS_ITEM", 2, minimo=1)
MAX_ABAS_PARALELAS = env_int("MAX_ABAS_PARALELAS", 10, minimo=1)
MAX_PAGINAS = env_int("MAX_PAGINAS", 0, minimo=0)  # 0 = sem limite

TIMEOUT_CURTO = 5_000
TIMEOUT_ITEM = env_int("TIMEOUT_ITEM", 3_000, minimo=500)
TIMEOUT_MEDIO = 30_000
TIMEOUT_LONGO = 60_000
TIMEOUT_NAVEGACAO = 90_000
ESPERA_APOS_APLICAR_FILTRO_MS = env_int("ESPERA_APOS_APLICAR_FILTRO_MS", 2_000, minimo=0)

URL_MERCADO_PAGO = "https://www.mercadopago.com.br/banking/balance/movements#from-section=balance"
PERFIL_AUTOMACAO = Path(os.getenv("PERFIL_AUTOMACAO", "perfil_google"))
PASTA_BASE_RELATORIOS = Path(os.getenv("PASTA_BASE_RELATORIOS", "output"))

CATEGORIAS_SEM_DETALHE = {"bônus por envio", "dinheiro recebido"}

# Movimentos operacionais que aparecem no extrato, mas não são vendas a exportar.
# O Mercado Pago pode renderizar esses itens sem cabeçalho de data durante a atualização da lista.
TERMOS_MOVIMENTO_IGNORADO = (
    "debito por divida",
    "devolucoes e reclamacoes",
    "debt-recovery",
    "poll-recovery",
    "detail: debt",
    "pp_recovery",
    "debt description",
    "debt, description",
)

MESES = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Março",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro",
}

MESES_POR_NOME = {nome.lower(): numero for numero, nome in MESES.items()}

PADRAO_DATA_EXTENSO = re.compile(r"(\d{1,2}) de ([a-zç]+)", re.IGNORECASE)
PADRAO_MES_ANO = re.compile(
    r"(Janeiro|Fevereiro|Março|Abril|Maio|Junho|Julho|Agosto|Setembro|Outubro|Novembro|Dezembro)\s+(\d{4})",
    re.IGNORECASE,
)
PADRAO_VALOR_RS = re.compile(r"R\$\s*([+-]?\s*[\d\.]+)\s*,\s*(\d{2})", re.IGNORECASE | re.DOTALL)
PADRAO_VALOR_SOLTO = re.compile(r"^[+-]?\s*[\d\.]+\s*,\s*\d{2}$")
PADRAO_HORA = re.compile(r"\b\d{1,2}h\d{2}\b")

SELETORES_LINHAS_MOVIMENTO = (
    "li:has-text('R$')",
    "div[role='button']:has-text('R$')",
    "a:has-text('R$')",
)


# =============================================================================
# EXCEÇÕES
# =============================================================================

class RoboMercadoPagoErro(RuntimeError):
    """Erro esperado do robô, com mensagem segura para o operador."""


class ErroCalendario(RoboMercadoPagoErro):
    """Falha ao ler ou manipular o calendário."""


class ErroExtracao(RoboMercadoPagoErro):
    """Falha ao extrair dados da tela."""


class ErroNavegacao(RoboMercadoPagoErro):
    """Falha ao navegar pelo Mercado Pago."""


ElementoMovimento = Locator | ElementHandle


# =============================================================================
# MODELOS
# =============================================================================

@dataclass(frozen=True)
class Periodo:
    inicio: datetime
    fim: datetime

    @property
    def mes_ano(self) -> str:
        return f"{MESES[self.inicio.month]} {self.inicio.year}"


@dataclass(frozen=True)
class Venda:
    data_venda: str
    numero_venda: str
    categoria: str
    valor_liquido: str
    valor_bruto: str = ""


@dataclass(frozen=True)
class ItemExtrato:
    data_venda: str
    categoria: str
    valor: str
    hora: str
    texto_completo: str


@dataclass(frozen=True)
class ErroItem:
    data_venda: str
    categoria: str
    valor: str
    hora: str
    motivo: str


@dataclass(frozen=True)
class ArquivosExecucao:
    planilha: Path
    log_erros: Path


@dataclass(frozen=True)
class ResultadoProcessamento:
    vendas: list[Venda]
    erros: list[ErroItem]
    total_itens: int


@dataclass(frozen=True)
class MovimentoCandidato:
    indice_original: int
    elemento: ElementHandle
    texto: str
    x: float = 0.0
    y: float = 0.0
    width: float = 0.0
    height: float = 0.0


# =============================================================================
# LOGGING
# =============================================================================

def log_info(msg: str) -> None:
    console.log(f"[blue]{msg}[/blue]")


def log_ok(msg: str) -> None:
    console.log(f"[green]{msg}[/green]")


def log_warn(msg: str) -> None:
    console.log(f"[yellow]{msg}[/yellow]")


def log_erro(msg: str) -> None:
    console.log(f"[bold red]{msg}[/bold red]")


def log_debug(msg: str) -> None:
    if MODO_DETALHADO:
        console.log(f"[dim]{msg}[/dim]")


# =============================================================================
# DATAS E PERÍODO
# =============================================================================

def ler_opcao_periodo() -> str:
    console.print("=" * 60)
    console.print("Escolha o tipo de período:")
    console.print("1 - Mês inteiro")
    console.print("2 - Ontem / pendente")
    console.print("3 - Data específica")
    console.print("4 - Período específico")
    console.print("=" * 60)

    while True:
        opcao = input("Digite 1, 2, 3 ou 4: ").strip()
        if opcao in {"1", "2", "3", "4"}:
            return opcao
        console.print("[yellow]Opção inválida.[/yellow]")


def ler_data_usuario(mensagem: str) -> datetime:
    while True:
        data_input = input(mensagem).strip()
        try:
            return datetime.strptime(data_input, "%d/%m/%Y")
        except ValueError:
            console.print("[yellow]Data inválida. Exemplo correto: 15/05/2026[/yellow]")


def ultimo_dia_mes_anterior(data_referencia: datetime) -> datetime:
    primeiro_dia_mes_atual = data_referencia.replace(day=1)
    return primeiro_dia_mes_atual - timedelta(days=1)


def calcular_periodo_mes_inteiro(hoje: datetime) -> Periodo:
    usar_mes_anterior = (
        hoje.day == 1
        or hoje.weekday() in {5, 6}
        or (hoje.weekday() == 0 and hoje.day <= 3)
    )

    fim = ultimo_dia_mes_anterior(hoje) if usar_mes_anterior else hoje - timedelta(days=1)
    inicio = fim.replace(day=1)
    return Periodo(inicio=inicio, fim=fim)


def calcular_periodo_ontem(hoje: datetime) -> Periodo:
    if hoje.weekday() == 0:
        inicio = hoje - timedelta(days=3)
        fim = hoje - timedelta(days=1)
    else:
        inicio = hoje - timedelta(days=1)
        fim = inicio

    return Periodo(inicio=inicio, fim=fim)


def calcular_periodo_data_especifica() -> Periodo:
    data = ler_data_usuario("Digite a data (DD/MM/AAAA): ")
    return Periodo(inicio=data, fim=data)


def calcular_periodo_especifico() -> Periodo:
    while True:
        inicio = ler_data_usuario("Digite a data inicial (DD/MM/AAAA): ")
        fim = ler_data_usuario("Digite a data final (DD/MM/AAAA): ")

        if fim < inicio:
            console.print("[yellow]A data final não pode ser menor que a data inicial.[/yellow]")
            continue

        if inicio.month != fim.month or inicio.year != fim.year:
            console.print(
                "[yellow]Este robô seleciona períodos dentro do mesmo mês. "
                "Digite um intervalo dentro do mesmo mês.[/yellow]"
            )
            continue

        return Periodo(inicio=inicio, fim=fim)


def calcular_periodo() -> Periodo:
    hoje = datetime.today()
    opcao = ler_opcao_periodo()

    if opcao == "1":
        periodo = calcular_periodo_mes_inteiro(hoje)
    elif opcao == "2":
        periodo = calcular_periodo_ontem(hoje)
    elif opcao == "3":
        periodo = calcular_periodo_data_especifica()
    else:
        periodo = calcular_periodo_especifico()

    console.print("=" * 60)
    console.print(
        f"Período selecionado: "
        f"{periodo.inicio.strftime('%d/%m/%Y')} até {periodo.fim.strftime('%d/%m/%Y')}"
    )
    console.print("=" * 60)

    return periodo


def converter_data_extenso(data_texto: str, ano_base: int) -> str:
    match = PADRAO_DATA_EXTENSO.search(data_texto or "")
    if not match:
        return data_texto

    dia = int(match.group(1))
    mes_nome = match.group(2).lower()
    mes = MESES_POR_NOME.get(mes_nome)

    if mes is None:
        raise ErroExtracao(f"Mês não reconhecido na data: {data_texto}")

    return f"{dia:02d}/{mes:02d}/{ano_base}"


def extrair_data_extenso_do_texto(texto: str) -> str:
    match = PADRAO_DATA_EXTENSO.search(texto or "")
    return match.group(0) if match else ""


# =============================================================================
# ARQUIVOS
# =============================================================================

def definir_arquivos_execucao(periodo: Periodo) -> ArquivosExecucao:
    pasta_ano = PASTA_BASE_RELATORIOS / str(periodo.inicio.year)
    pasta_mes = pasta_ano / f"{periodo.inicio.month:02d}.{periodo.inicio.year}"
    pasta_mes.mkdir(parents=True, exist_ok=True)

    nome_planilha = (
        f"MP - EXTRATO "
        f"{periodo.inicio.day:02d}.{periodo.inicio.month:02d}.{periodo.inicio.year} - "
        f"{periodo.fim.day:02d}.{periodo.fim.month:02d}.{periodo.fim.year}.xlsx"
    )

    planilha = pasta_mes / nome_planilha
    log_erros = planilha.with_suffix(".txt")

    return ArquivosExecucao(planilha=planilha, log_erros=log_erros)


# =============================================================================
# DINHEIRO E EXCEL
# =============================================================================

def valor_para_decimal(valor: object) -> Decimal | None:
    if valor is None:
        return None

    if isinstance(valor, Decimal):
        return valor.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    texto = str(valor).strip()
    if not texto:
        return None

    texto = texto.replace("R$", "").replace(" ", "")

    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")

    texto = re.sub(r"[^\d\-.]", "", texto)
    if texto in {"", "-", "."}:
        return None

    try:
        return Decimal(texto).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except InvalidOperation as erro:
        raise ErroExtracao(f"Valor monetário inválido: {valor!r}") from erro


def decimal_para_float(valor: Decimal | None) -> float | None:
    if valor is None:
        return None
    return float(valor)


def chave_valor(valor: object) -> str:
    decimal = valor_para_decimal(valor)
    return "" if decimal is None else f"{decimal:.2f}"


def normalizar_texto_planilha(valor: object) -> str:
    return "" if valor is None else str(valor).strip()


def preparar_planilha(caminho: Path) -> tuple[Workbook, Worksheet]:
    caminho.parent.mkdir(parents=True, exist_ok=True)

    if caminho.exists():
        wb = load_workbook(caminho)
        ws = wb.active
        garantir_cabecalho_planilha(ws)
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimentos"
    garantir_cabecalho_planilha(ws)
    return wb, ws


def garantir_cabecalho_planilha(ws: Worksheet) -> None:
    cabecalhos = ["DATA", "#VENDA", "CATEGORIA", "VALOR LÍQUIDO", "VALOR BRUTO"]
    larguras = [15, 22, 35, 15, 15]

    for indice, cabecalho in enumerate(cabecalhos, start=1):
        celula = ws.cell(row=1, column=indice)
        if not celula.value:
            celula.value = cabecalho

        fonte = copy(celula.font)
        fonte.bold = True
        celula.font = fonte
        ws.column_dimensions[celula.column_letter].width = larguras[indice - 1]

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:E1"


def chave_linha_planilha(linha: tuple[object, ...]) -> tuple[str, str, str, str, str]:
    valores = list(linha[:5]) + [None] * max(0, 5 - len(linha[:5]))
    return (
        normalizar_texto_planilha(valores[0]),
        normalizar_texto_planilha(valores[1]),
        normalizar_texto_planilha(valores[2]),
        chave_valor(valores[3]),
        chave_valor(valores[4]),
    )


def obter_registros_existentes(ws: Worksheet) -> set[tuple[str, str, str, str, str]]:
    registros: set[tuple[str, str, str, str, str]] = set()

    for linha in ws.iter_rows(min_row=2, values_only=True):
        if not any(valor is not None and str(valor).strip() for valor in linha[:5]):
            continue
        registros.add(chave_linha_planilha(linha))

    return registros


def chave_venda(venda: Venda) -> tuple[str, str, str, str, str]:
    return (
        normalizar_texto_planilha(venda.data_venda),
        normalizar_texto_planilha(venda.numero_venda),
        normalizar_texto_planilha(venda.categoria),
        chave_valor(venda.valor_liquido),
        chave_valor(venda.valor_bruto),
    )


def adicionar_venda_na_planilha(ws: Worksheet, venda: Venda) -> None:
    valor_liquido = valor_para_decimal(venda.valor_liquido)
    valor_bruto = valor_para_decimal(venda.valor_bruto)

    if valor_liquido is None:
        raise ErroExtracao(f"Venda sem valor líquido válido: {venda}")

    linha = ws.max_row + 1
    ws.cell(row=linha, column=1, value=venda.data_venda)
    ws.cell(row=linha, column=2, value=venda.numero_venda)
    ws.cell(row=linha, column=3, value=venda.categoria)

    celula_liquido = ws.cell(row=linha, column=4, value=decimal_para_float(valor_liquido))
    celula_liquido.number_format = "R$ #,##0.00"

    celula_bruto = ws.cell(row=linha, column=5)
    if valor_bruto is not None:
        celula_bruto.value = decimal_para_float(valor_bruto)
        celula_bruto.number_format = "R$ #,##0.00"
    else:
        celula_bruto.value = ""


def exportar_vendas_para_planilha(vendas: list[Venda], caminho: Path) -> int:
    if not vendas:
        log_warn("Nenhuma venda para exportar.")
        return 0

    wb, ws = preparar_planilha(caminho)
    registros_existentes = obter_registros_existentes(ws)
    total_exportado = 0

    for venda in vendas:
        chave = chave_venda(venda)
        if chave in registros_existentes:
            log_debug(f"Linha já existe na planilha. Ignorando: {chave}")
            continue

        adicionar_venda_na_planilha(ws, venda)
        registros_existentes.add(chave)
        total_exportado += 1

    wb.save(caminho)
    console.print(f"[green]✓[/green] {total_exportado} venda(s) exportada(s) para Excel.")
    return total_exportado


# =============================================================================
# PLAYWRIGHT - AÇÕES BÁSICAS
# =============================================================================

async def aguardar_carregamento_basico(pagina: Page, timeout: int = TIMEOUT_LONGO) -> None:
    try:
        await pagina.wait_for_load_state("domcontentloaded", timeout=timeout)
    except PlaywrightTimeoutError:
        log_debug("Timeout aguardando domcontentloaded. Continuando com verificações de tela.")


async def aguardar_estabilizacao_visual(pagina: Page, ms: int = 250) -> None:
    # Pequena espera deliberada para transições de calendário/filtros.
    # O restante do código usa esperas por elementos visíveis.
    await pagina.wait_for_timeout(ms)


async def obter_texto_corpo_seguro(pagina: Page, timeout: int = TIMEOUT_CURTO) -> str:
    try:
        return await pagina.locator("body").inner_text(timeout=timeout)
    except Exception:
        return ""


async def aguardar_atualizacao_pos_filtro(pagina: Page, texto_antes: str = "") -> None:
    await aguardar_carregamento_basico(pagina)

    if texto_antes:
        try:
            await pagina.wait_for_function(
                "textoAntes => document.body && document.body.innerText !== textoAntes",
                arg=texto_antes,
                timeout=10_000,
            )
        except PlaywrightTimeoutError:
            log_debug("Texto da página não mudou rapidamente após aplicar filtro; aguardando estabilização fixa.")

    await aguardar_estabilizacao_visual(pagina, ESPERA_APOS_APLICAR_FILTRO_MS)


async def clicar_texto(container: Page | Locator, texto: str, exact: bool = True, timeout: int = TIMEOUT_MEDIO) -> None:
    item = container.get_by_text(texto, exact=exact).first
    await item.wait_for(state="visible", timeout=timeout)
    await item.click()


async def clicar_botao(container: Page | Locator, nome: str | re.Pattern[str], timeout: int = TIMEOUT_MEDIO) -> None:
    botao = container.get_by_role("button", name=nome).first
    await botao.wait_for(state="visible", timeout=timeout)
    await botao.click()


async def esperar_texto_visivel(
    pagina: Page,
    texto: str,
    exact: bool = True,
    timeout: int = TIMEOUT_MEDIO,
) -> Locator:
    elemento = pagina.get_by_text(texto, exact=exact).first
    await elemento.wait_for(state="visible", timeout=timeout)
    return elemento


async def existe_texto(pagina: Page, texto: str, exact: bool = True, timeout: int = 3_000) -> bool:
    try:
        await pagina.get_by_text(texto, exact=exact).first.wait_for(state="visible", timeout=timeout)
        return True
    except PlaywrightTimeoutError:
        return False


async def abrir_navegador(playwright) -> BrowserContext:
    return await playwright.chromium.launch_persistent_context(
        user_data_dir=str(PERFIL_AUTOMACAO),
        channel="chrome",
        headless=False,
        accept_downloads=True,
        args=[
            "--start-minimized",
            "--disable-features=Translate",
            "--disable-translate",
            "--no-first-run",
            "--no-default-browser-check",
        ],
        no_viewport=True,
    )


async def fechar_abas_em_branco(navegador: BrowserContext, pagina_principal: Page) -> None:
    for aba in navegador.pages:
        if aba.url == "about:blank" and aba != pagina_principal:
            await fechar_guia_com_segurança(aba)


async def fechar_guia_com_segurança(pagina: Optional[Page]) -> None:
    try:
        if pagina and not pagina.is_closed():
            await pagina.close()
    except Exception as erro:
        log_debug(f"Falha ao fechar guia: {erro}")


# =============================================================================
# CALENDÁRIO MERCADO PAGO
# =============================================================================

async def obter_mes_ano_calendario(calendario: Locator) -> tuple[int, int]:
    texto = await calendario.inner_text(timeout=10_000)
    match = PADRAO_MES_ANO.search(texto)

    if not match:
        raise ErroCalendario(f"Não consegui identificar o mês/ano do calendário. Texto capturado:\n{texto[:1000]}")

    mes = MESES_POR_NOME[match.group(1).lower()]
    ano = int(match.group(2))
    return mes, ano


def diferenca_meses(mes_atual: int, ano_atual: int, mes_alvo: int, ano_alvo: int) -> int:
    return (ano_alvo - ano_atual) * 12 + (mes_alvo - mes_atual)


async def aguardar_mudanca_mes(calendario: Locator, mes_ano_anterior: tuple[int, int]) -> None:
    for _ in range(30):
        try:
            if await obter_mes_ano_calendario(calendario) != mes_ano_anterior:
                return
        except ErroCalendario:
            pass
        await asyncio.sleep(0.1)


async def ajustar_mes_calendario(pagina: Page, calendario: Locator, data_alvo: datetime) -> None:
    for _ in range(24):
        mes_atual, ano_atual = await obter_mes_ano_calendario(calendario)
        diferenca = diferenca_meses(mes_atual, ano_atual, data_alvo.month, data_alvo.year)

        if diferenca == 0:
            log_ok(f"Calendário ajustado para {MESES[data_alvo.month]} {data_alvo.year}.")
            return

        botoes = calendario.locator("button")
        botao = botoes.nth(0) if diferenca < 0 else botoes.nth(1)
        acao = "Voltando" if diferenca < 0 else "Avançando"
        log_info(f"{acao} um mês no calendário...")

        await botao.click()
        await aguardar_mudanca_mes(calendario, (mes_atual, ano_atual))
        await aguardar_estabilizacao_visual(pagina, 150)

    raise ErroCalendario(f"Não consegui ajustar o calendário para {MESES[data_alvo.month]} {data_alvo.year}.")


async def clicar_data_calendario(calendario: Locator, data: datetime) -> None:
    primeiro_dia_mes = data.replace(day=1)
    indice_primeiro_dia = (primeiro_dia_mes.weekday() + 1) % 7
    indice_alvo = indice_primeiro_dia + data.day - 1

    botoes = calendario.locator("button")
    candidatos: list[dict[str, object]] = []

    for indice in range(await botoes.count()):
        botao = botoes.nth(indice)

        try:
            texto = (await botao.inner_text(timeout=1_000)).strip()
        except PlaywrightTimeoutError:
            continue

        if not re.fullmatch(r"\d{1,2}", texto):
            continue

        box = await botao.bounding_box()
        if not box:
            continue

        candidatos.append({"botao": botao, "texto": texto, "x": box["x"], "y": box["y"]})

    if not candidatos:
        raise ErroCalendario("Nenhum botão de dia foi encontrado no calendário.")

    candidatos_ordenados = sorted(candidatos, key=lambda item: (float(item["y"]), float(item["x"])))

    if indice_alvo >= len(candidatos_ordenados):
        dias_visiveis = [str(item["texto"]) for item in candidatos_ordenados]
        raise ErroCalendario(
            f"Índice do dia fora da grade do calendário. "
            f"Data alvo: {data.strftime('%d/%m/%Y')}. Dias visíveis: {dias_visiveis}"
        )

    candidato = candidatos_ordenados[indice_alvo]
    texto_candidato = str(candidato["texto"])

    if int(texto_candidato) != data.day:
        dias_visiveis = [str(item["texto"]) for item in candidatos_ordenados]
        raise ErroCalendario(
            f"O botão calculado não bate com o dia esperado. "
            f"Esperado: {data.day}. Encontrado: {texto_candidato}. Dias visíveis: {dias_visiveis}"
        )

    log_info(f"Clicando na data correta: {data.strftime('%d/%m/%Y')}")
    botao_alvo = candidato["botao"]
    if not isinstance(botao_alvo, Locator):
        raise ErroCalendario("Candidato de data inválido.")

    await botao_alvo.click(force=True)


# =============================================================================
# EXTRAÇÃO
# =============================================================================

def normalizar_linhas(texto: str) -> list[str]:
    return [linha.strip() for linha in texto.splitlines() if linha.strip()]


def normalizar_para_comparacao(texto: str) -> str:
    texto_normalizado = unicodedata.normalize("NFKD", texto or "")
    texto_sem_acentos = "".join(
        caractere for caractere in texto_normalizado
        if not unicodedata.combining(caractere)
    )
    return re.sub(r"\s+", " ", texto_sem_acentos).strip().lower()


def movimento_operacional_ignorado(texto_item: str, categoria: str = "") -> bool:
    texto_comparacao = normalizar_para_comparacao(f"{categoria}\n{texto_item}")
    return any(termo in texto_comparacao for termo in TERMOS_MOVIMENTO_IGNORADO)


def linha_irrelevante_para_categoria(linha: str) -> bool:
    if PADRAO_DATA_EXTENSO.fullmatch(linha):
        return True
    if linha in {"R$", ","}:
        return True
    if re.fullmatch(r"\d+", linha):
        return True
    if PADRAO_HORA.search(linha):
        return True
    if "R$" in linha:
        return True
    if PADRAO_VALOR_SOLTO.fullmatch(linha):
        return True
    return False


def identificar_categoria(linhas: list[str]) -> str:
    for linha in linhas:
        if linha_irrelevante_para_categoria(linha):
            continue
        return linha

    return "Não identificado"


def identificar_valor_item(texto: str) -> str:
    match = PADRAO_VALOR_RS.search(texto)
    if not match:
        raise ErroExtracao(f"Valor não encontrado no item:\n{texto}")

    reais = match.group(1).replace(" ", "")
    centavos = match.group(2)
    return f"{reais},{centavos}"


def identificar_hora_item(texto: str) -> str:
    match = PADRAO_HORA.search(texto)
    return match.group(0) if match else ""


async def obter_texto_elemento(item: ElementoMovimento, timeout: int = TIMEOUT_ITEM) -> str:
    if isinstance(item, Locator):
        return (await item.inner_text(timeout=timeout)).strip()

    try:
        texto = await asyncio.wait_for(item.inner_text(), timeout=timeout / 1000)
    except asyncio.TimeoutError as erro:
        raise ErroExtracao(f"Timeout ao ler texto do movimento após {timeout} ms.") from erro

    return texto.strip()


async def extrair_dados_item(
    item: ElementoMovimento,
    data_atual: str,
    texto_item: str | None = None,
) -> ItemExtrato:
    texto = texto_item if texto_item is not None else await obter_texto_elemento(item)
    texto = texto.strip()
    linhas = normalizar_linhas(texto)

    categoria = identificar_categoria(linhas)
    valor = identificar_valor_item(texto)
    hora = identificar_hora_item(texto)

    if not data_atual:
        raise ErroExtracao(f"Data do item não encontrada. Texto do item:\n{texto[:1000]}")

    return ItemExtrato(
        data_venda=data_atual,
        categoria=categoria,
        valor=valor,
        hora=hora,
        texto_completo=texto,
    )


async def extrair_numero_venda(pagina: Page) -> str:
    texto = await pagina.locator("body").inner_text()

    padroes = [
        r"Venda\s+(2000\d+)",
        r"#\s*(2000\d+)",
        r"\b(2000\d{8,})\b",
    ]

    for padrao in padroes:
        match = re.search(padrao, texto, re.IGNORECASE)
        if match:
            return match.group(1)

    return ""


async def extrair_valor_bruto(pagina: Page) -> str:
    texto = await pagina.locator("body").inner_text()

    padroes_rotulados = [
        r"Valor da venda\s*R\$\s*([\d\.]+)\s*,\s*(\d{2})",
        r"Valor bruto\s*R\$\s*([\d\.]+)\s*,\s*(\d{2})",
        r"Total pago pelo comprador\s*R\$\s*([\d\.]+)\s*,\s*(\d{2})",
        r"Pagamento recebido\s*R\$\s*([\d\.]+)\s*,\s*(\d{2})",
    ]

    for padrao in padroes_rotulados:
        match = re.search(padrao, texto, re.IGNORECASE | re.DOTALL)
        if match:
            return f"{match.group(1)},{match.group(2)}"

    match = PADRAO_VALOR_RS.search(texto)
    if match:
        return f"{match.group(1).replace(' ', '')},{match.group(2)}"

    match_sem_centavos = re.search(r"R\$\s*([\d\.]+)(?!\s*,)", texto, re.IGNORECASE)
    if match_sem_centavos:
        return f"{match_sem_centavos.group(1)},00"

    return ""


async def obter_data_mais_proxima(pagina: Page, item: ElementoMovimento) -> str:
    item_box = await item.bounding_box()
    if not item_box:
        return ""

    datas = pagina.locator("text=/\\d{1,2} de [a-zç]+/i")
    melhor_data = ""

    for indice in range(await datas.count()):
        data_locator = datas.nth(indice)
        data_box = await data_locator.bounding_box()

        if data_box and data_box["y"] <= item_box["y"]:
            try:
                melhor_data = (await data_locator.inner_text(timeout=1_000)).strip()
            except PlaywrightTimeoutError:
                continue

    return melhor_data


# =============================================================================
# MERCADO PAGO - FLUXO DE TELA
# =============================================================================

async def abrir_mercado_pago(pagina: Page) -> None:
    log_info("Abrindo Mercado Pago...")
    await pagina.goto(URL_MERCADO_PAGO, wait_until="domcontentloaded", timeout=TIMEOUT_NAVEGACAO)

    try:
        await pagina.get_by_text("Período", exact=False).first.wait_for(state="visible", timeout=TIMEOUT_LONGO)
    except PlaywrightTimeoutError as erro:
        corpo = (await pagina.locator("body").inner_text())[:3000]
        console.print("[red]Não encontrei 'Período'. A página pode estar em login, validação ou carregamento parcial.[/red]")
        console.print(f"URL atual: {pagina.url}")
        console.print(corpo)
        raise ErroNavegacao("Mercado Pago não carregou a tela de movimentos.") from erro


async def abrir_menu_periodo(pagina: Page) -> None:
    for tentativa in range(1, 4):
        try:
            log_info(f"Tentativa Período {tentativa}/3")
            await clicar_texto(pagina, "Período", exact=False, timeout=10_000)
            await pagina.get_by_text("Outro período", exact=True).first.wait_for(state="visible", timeout=5_000)
            log_ok("Menu de período abriu.")
            return
        except Exception as erro:
            log_warn(f"Não abriu o menu de período: {erro}")
            if tentativa == 3:
                raise ErroNavegacao("Não consegui abrir o menu de Período após 3 tentativas.") from erro
            await aguardar_estabilizacao_visual(pagina, 500)


async def selecionar_periodo(pagina: Page, periodo: Periodo) -> None:
    await abrir_menu_periodo(pagina)
    await clicar_texto(pagina, "Outro período", exact=True)

    calendario = pagina.locator("div:has-text('Aplicar')").last
    await calendario.wait_for(state="visible", timeout=TIMEOUT_MEDIO)

    await ajustar_mes_calendario(pagina, calendario, periodo.inicio)
    await clicar_data_calendario(calendario, periodo.inicio)
    await aguardar_estabilizacao_visual(pagina, 250)

    await clicar_data_calendario(calendario, periodo.fim)
    await aguardar_estabilizacao_visual(pagina, 250)

    texto_antes = await obter_texto_corpo_seguro(pagina)
    await clicar_botao(calendario, "Aplicar")
    await aguardar_atualizacao_pos_filtro(pagina, texto_antes)


async def filtrar_todas_entradas(pagina: Page) -> None:
    log_info("Aguardando botão Transações aparecer...")
    await pagina.get_by_text("Transações", exact=False).first.wait_for(state="visible", timeout=TIMEOUT_LONGO)
    await clicar_texto(pagina, "Transações", exact=False)

    painel_transacoes = pagina.locator(
        "div:has-text('Todas as transações'):has-text('Todas as entradas'):has-text('Aplicar')"
    ).last
    await painel_transacoes.wait_for(state="visible", timeout=TIMEOUT_MEDIO)

    await clicar_texto(painel_transacoes, "Todas as entradas", exact=True)

    texto_antes = await obter_texto_corpo_seguro(pagina)
    await clicar_botao(painel_transacoes, "Aplicar")
    await aguardar_atualizacao_pos_filtro(pagina, texto_antes)


async def aguardar_lista_movimentos(pagina: Page) -> None:
    await pagina.locator("text=/\\d+ de [a-zç]+/i").first.wait_for(state="visible", timeout=TIMEOUT_LONGO)
    await pagina.locator(", ".join(SELETORES_LINHAS_MOVIMENTO)).first.wait_for(state="visible", timeout=TIMEOUT_LONGO)
    await aguardar_estabilizacao_visual(pagina, 500)


async def clicar_atividade_vinculada(pagina: Page, aguardar_url: bool = True) -> None:
    log_info("Tentando clicar em Atividade vinculada...")
    await aguardar_carregamento_basico(pagina)

    for tentativa in range(1, 3):
        try:
            log_info(f"Tentativa Atividade vinculada {tentativa}/2")

            texto = pagina.get_by_text("Atividade vinculada", exact=True).first
            await texto.wait_for(state="visible", timeout=3_000)

            try:
                await texto.scroll_into_view_if_needed(timeout=10_000)
            except Exception as erro:
                log_debug(f"Scroll até Atividade vinculada falhou, tentando clique mesmo assim: {erro}")

            await texto.click(force=True, timeout=5_000)

            if not aguardar_url:
                return

            try:
                await pagina.wait_for_url("**/activities/detail/**", timeout=3_000)
            except PlaywrightTimeoutError:
                pass

            if "/activities/detail/" in pagina.url:
                log_ok("Atividade vinculada aberta.")
                return

            await aguardar_carregamento_basico(pagina, timeout=10_000)

        except Exception as erro:
            log_warn(f"Falhou ao clicar em Atividade vinculada na tentativa {tentativa}: {erro}")
            await aguardar_estabilizacao_visual(pagina, 500)

    raise ErroNavegacao("Não consegui abrir 'Atividade vinculada' após várias tentativas.")


async def clicar_descricao_venda(pagina: Page) -> None:
    """Abre o card da venda após entrar em Atividade vinculada.

    Após atualização do Mercado Pago, a tela passou a exibir um bloco intermediário
    com "Descrição da venda" e o nome do produto. O robô precisa clicar nesse bloco
    antes de procurar "Conferir detalhes".
    """
    log_info("Tentando clicar na descrição da venda...")
    await aguardar_carregamento_basico(pagina)

    if await existe_texto(pagina, "Conferir detalhes", exact=True, timeout=1_500):
        log_debug("'Conferir detalhes' já está visível; pulando clique na descrição da venda.")
        return

    erro_final: Exception | None = None

    for tentativa in range(1, 4):
        try:
            log_info(f"Tentativa Descrição da venda {tentativa}/3")

            descricao = pagina.get_by_text("Descrição da venda", exact=True).first
            await descricao.wait_for(state="visible", timeout=TIMEOUT_MEDIO)
            await descricao.scroll_into_view_if_needed(timeout=10_000)

            candidatos = [
                descricao.locator("xpath=ancestor::*[self::a or self::button or @role='button'][1]"),
                descricao.locator("xpath=ancestor::*[contains(@class, 'andes-card')][1]"),
                descricao.locator("xpath=ancestor::div[1]"),
                descricao.locator("xpath=ancestor::div[2]"),
                descricao.locator("xpath=ancestor::div[3]"),
                descricao,
            ]

            for candidato in candidatos:
                try:
                    if await candidato.count() == 0:
                        continue

                    alvo = candidato.first
                    caixa = await alvo.bounding_box()
                    if not caixa:
                        continue

                    await alvo.click(force=True, timeout=5_000)
                    await aguardar_carregamento_basico(pagina, timeout=10_000)
                    await aguardar_estabilizacao_visual(pagina, 800)

                    if await existe_texto(pagina, "Conferir detalhes", exact=True, timeout=4_000):
                        log_ok("Descrição da venda aberta.")
                        return

                    if await existe_texto(pagina, "Ver detalhes da venda", exact=True, timeout=2_000):
                        log_ok("Descrição da venda aberta.")
                        return

                except Exception as erro_candidato:
                    erro_final = erro_candidato
                    log_debug(f"Candidato de clique na descrição da venda falhou: {erro_candidato}")

            # Fallback: clique no centro visual do texto. Em alguns layouts, o listener
            # fica no card pai e o evento sobe por bubbling.
            caixa_descricao = await descricao.bounding_box()
            if caixa_descricao:
                await pagina.mouse.click(
                    caixa_descricao["x"] + caixa_descricao["width"] / 2,
                    caixa_descricao["y"] + caixa_descricao["height"] / 2,
                )
                await aguardar_carregamento_basico(pagina, timeout=10_000)
                await aguardar_estabilizacao_visual(pagina, 800)

                if await existe_texto(pagina, "Conferir detalhes", exact=True, timeout=4_000):
                    log_ok("Descrição da venda aberta.")
                    return

                if await existe_texto(pagina, "Ver detalhes da venda", exact=True, timeout=2_000):
                    log_ok("Descrição da venda aberta.")
                    return

        except Exception as erro:
            erro_final = erro
            log_warn(f"Falhou ao clicar em Descrição da venda na tentativa {tentativa}: {erro}")
            await aguardar_estabilizacao_visual(pagina, 500)

    raise ErroNavegacao(
        "Não consegui abrir o bloco 'Descrição da venda' após clicar em Atividade vinculada."
    ) from erro_final


async def clicar_ver_detalhes_da_venda(pagina: Page) -> None:
    """Avança da descrição/modal da venda para a tela final de detalhes."""
    await aguardar_carregamento_basico(pagina)
    await aguardar_estabilizacao_visual(pagina, 500)

    # Layout antigo/intermediário: antes do botão final havia "Conferir detalhes".
    if await existe_texto(pagina, "Conferir detalhes", exact=True, timeout=2_000):
        conferir = pagina.get_by_text("Conferir detalhes", exact=True).first
        await conferir.scroll_into_view_if_needed(timeout=10_000)
        await conferir.click(force=True, timeout=10_000)
        await aguardar_carregamento_basico(pagina, timeout=10_000)
        await aguardar_estabilizacao_visual(pagina, 500)

    erro_final: Exception | None = None

    for tentativa in range(1, 4):
        try:
            log_info(f"Tentativa Ver detalhes da venda {tentativa}/3")

            candidatos = [
                pagina.get_by_role("button", name=re.compile(r"Ver detalhes da venda", re.IGNORECASE)).first,
                pagina.get_by_text("Ver detalhes da venda", exact=True).first,
                pagina.locator("text=/Ver detalhes da venda/i").first,
            ]

            for candidato in candidatos:
                try:
                    await candidato.wait_for(state="visible", timeout=8_000)
                    await candidato.scroll_into_view_if_needed(timeout=10_000)
                    await candidato.click(force=True, timeout=10_000)
                    await aguardar_carregamento_basico(pagina, timeout=15_000)
                    await aguardar_estabilizacao_visual(pagina, 700)
                    log_ok("Botão 'Ver detalhes da venda' acionado.")
                    return
                except Exception as erro_candidato:
                    erro_final = erro_candidato
                    log_debug(f"Candidato 'Ver detalhes da venda' falhou: {erro_candidato}")

        except Exception as erro:
            erro_final = erro
            log_warn(f"Falhou ao clicar em Ver detalhes da venda na tentativa {tentativa}: {erro}")
            await aguardar_estabilizacao_visual(pagina, 500)

    raise ErroNavegacao("Não consegui clicar em 'Ver detalhes da venda'.") from erro_final


async def extrair_venda_da_guia(item_extraido: ItemExtrato, guia: Page) -> Venda:
    await aguardar_carregamento_basico(guia)

    if not await existe_texto(guia, "Atividade vinculada", timeout=5_000):
        raise ErroNavegacao("Atividade vinculada não encontrada.")

    await clicar_atividade_vinculada(guia)
    await clicar_descricao_venda(guia)
    await clicar_ver_detalhes_da_venda(guia)

    numero_venda = re.sub(r"\D", "", await extrair_numero_venda(guia) or "")
    if not numero_venda:
        raise ErroExtracao("Número da venda não encontrado na tela de detalhes.")

    valor_bruto = await extrair_valor_bruto(guia)

    return Venda(
        data_venda=item_extraido.data_venda,
        numero_venda=numero_venda,
        categoria=item_extraido.categoria,
        valor_liquido=item_extraido.valor,
        valor_bruto=valor_bruto,
    )


async def processar_guia_individual(
    item_extraido: ItemExtrato,
    nova_guia: Page,
) -> tuple[Optional[Venda], Optional[ErroItem]]:
    erro_final: Exception | None = None

    try:
        nova_guia.set_default_timeout(TIMEOUT_LONGO)

        for tentativa in range(1, MAX_TENTATIVAS_ITEM + 1):
            try:
                venda = await extrair_venda_da_guia(item_extraido, nova_guia)
                return venda, None
            except Exception as erro:
                erro_final = erro
                log_warn(
                    f"Falha ao detalhar item ({tentativa}/{MAX_TENTATIVAS_ITEM}): "
                    f"{item_extraido.data_venda} | {item_extraido.categoria} | R$ {item_extraido.valor} | {erro}"
                )

                if tentativa < MAX_TENTATIVAS_ITEM and not nova_guia.is_closed():
                    try:
                        await nova_guia.reload(wait_until="domcontentloaded", timeout=TIMEOUT_LONGO)
                    except Exception as reload_erro:
                        log_debug(f"Reload da guia falhou: {reload_erro}")

        return None, ErroItem(
            data_venda=item_extraido.data_venda,
            categoria=item_extraido.categoria,
            valor=item_extraido.valor,
            hora=item_extraido.hora,
            motivo=str(erro_final) if erro_final else "Falha desconhecida ao detalhar item.",
        )

    finally:
        await fechar_guia_com_segurança(nova_guia)


async def processar_lote_itens(
    navegador: BrowserContext,
    pagina: Page,
    lote: list[tuple[ItemExtrato, ElementoMovimento]],
) -> tuple[list[Venda], list[ErroItem]]:
    vendas: list[Venda] = []
    erros: list[ErroItem] = []
    guias_abertas: list[tuple[ItemExtrato, Page]] = []

    for item_extraido, item in lote:
        try:
            try:
                await item.scroll_into_view_if_needed(timeout=TIMEOUT_CURTO)
            except Exception as erro_scroll:
                log_debug(f"Scroll do item antes de abrir nova guia falhou: {erro_scroll}")

            async with navegador.expect_page(timeout=TIMEOUT_MEDIO) as nova_guia_info:
                await item.click(button="middle", timeout=TIMEOUT_CURTO)

            nova_guia = await nova_guia_info.value
            nova_guia.set_default_timeout(TIMEOUT_LONGO)
            guias_abertas.append((item_extraido, nova_guia))
            await pagina.bring_to_front()

        except Exception as erro:
            erros.append(
                ErroItem(
                    data_venda=item_extraido.data_venda,
                    categoria=item_extraido.categoria,
                    valor=item_extraido.valor,
                    hora=item_extraido.hora,
                    motivo=f"Falha ao abrir nova guia: {erro}",
                )
            )

    tarefas = [
        processar_guia_individual(item_extraido, nova_guia)
        for item_extraido, nova_guia in guias_abertas
    ]

    for resultado in await asyncio.gather(*tarefas, return_exceptions=True):
        if isinstance(resultado, Exception):
            erros.append(
                ErroItem(
                    data_venda="",
                    categoria="",
                    valor="",
                    hora="",
                    motivo=f"Erro inesperado no lote: {resultado}",
                )
            )
            continue

        venda, erro = resultado
        if venda:
            vendas.append(venda)
        if erro:
            erros.append(erro)

    await pagina.bring_to_front()
    return vendas, erros


# =============================================================================
# PROCESSAMENTO DE PÁGINAS
# =============================================================================

async def obter_linhas_movimento(pagina: Page) -> list[MovimentoCandidato]:
    """
    Captura um snapshot somente dos movimentos reais/visíveis.

    O Mercado Pago costuma manter containers, linhas antigas ou wrappers no DOM.
    Se o robô usar todos os `li:has-text('R$')`, ele pode clicar em um item
    adicional que não aparece como movimento real na tela. Por isso o snapshot
    agora filtra elementos invisíveis, wrappers com data no início e candidatos
    ancestrais/duplicados que representam a mesma movimentação.
    """
    for seletor in SELETORES_LINHAS_MOVIMENTO:
        linhas = pagina.locator(seletor)

        try:
            await linhas.first.wait_for(state="attached", timeout=TIMEOUT_CURTO)
        except PlaywrightTimeoutError:
            continue

        elementos = await linhas.element_handles()
        movimentos_relaxados: list[MovimentoCandidato] = []

        for indice, elemento in enumerate(elementos):
            try:
                if not await elemento.is_visible():
                    continue

                caixa = await elemento.bounding_box()
                if not caixa or caixa.get("width", 0) <= 0 or caixa.get("height", 0) <= 0:
                    continue

                texto = await obter_texto_elemento(elemento, timeout=TIMEOUT_ITEM)
            except Exception as erro:
                log_debug(f"Elemento instável ignorado no seletor {seletor}, índice {indice + 1}: {erro}")
                continue

            if item_deve_ser_ignorado(texto, descartar_container_data=False):
                continue

            movimentos_relaxados.append(
                MovimentoCandidato(
                    indice_original=indice,
                    elemento=elemento,
                    texto=texto,
                    x=float(caixa["x"]),
                    y=float(caixa["y"]),
                    width=float(caixa["width"]),
                    height=float(caixa["height"]),
                )
            )

        movimentos = await filtrar_candidatos_movimento(movimentos_relaxados)
        movimentos_sem_container_data = [
            movimento
            for movimento in movimentos
            if not item_deve_ser_ignorado(movimento.texto, descartar_container_data=True)
        ]

        # Caminho normal do layout atual: cada item fica separado do cabeçalho de data.
        # Se algum layout futuro embutir a data dentro do próprio item, o fallback evita
        # zerar a página inteira.
        if movimentos_sem_container_data:
            return movimentos_sem_container_data

        if movimentos:
            return movimentos

    return []


def texto_comeca_com_data(texto_item: str) -> bool:
    linhas = normalizar_linhas(texto_item)
    return bool(linhas and PADRAO_DATA_EXTENSO.fullmatch(linhas[0]))


def item_deve_ser_ignorado(texto_item: str, descartar_container_data: bool = True) -> bool:
    if "R$" not in texto_item:
        return True

    if descartar_container_data and texto_comeca_com_data(texto_item):
        # No layout atual do Mercado Pago, uma linha que começa com "26 de maio"
        # normalmente é um wrapper do bloco do dia, não o movimento clicável.
        # Processar esse wrapper causa item extra/duplicado.
        return True

    # Evita capturar containers grandes que agrupam vários movimentos.
    if texto_item.count("R$") > 2 and len(texto_item) > 500:
        return True

    return len(texto_item) > 1_500


def area_movimento(movimento: MovimentoCandidato) -> float:
    return max(movimento.width, 0.0) * max(movimento.height, 0.0)


def sobreposicao_vertical(a: MovimentoCandidato, b: MovimentoCandidato) -> float:
    topo = max(a.y, b.y)
    base = min(a.y + a.height, b.y + b.height)
    return max(0.0, base - topo)


def movimentos_sobrepostos(a: MovimentoCandidato, b: MovimentoCandidato) -> bool:
    menor_altura = max(min(a.height, b.height), 1.0)
    return sobreposicao_vertical(a, b) >= menor_altura * 0.60


def assinatura_movimento(texto_item: str) -> str:
    try:
        categoria = identificar_categoria(normalizar_linhas(texto_item))
    except Exception:
        categoria = ""

    try:
        valor = identificar_valor_item(texto_item)
    except Exception:
        valor = ""

    data = extrair_data_extenso_do_texto(texto_item)
    hora = identificar_hora_item(texto_item)

    return normalizar_para_comparacao(f"{data}|{categoria}|{valor}|{hora}")


def score_preferencia_movimento(movimento: MovimentoCandidato) -> tuple[int, int, int, float]:
    # Menor score = candidato mais provável de ser a linha real clicável.
    return (
        1 if texto_comeca_com_data(movimento.texto) else 0,
        movimento.texto.count("R$"),
        len(normalizar_linhas(movimento.texto)),
        area_movimento(movimento),
    )


async def elemento_contem_outro(pai: MovimentoCandidato, filho: MovimentoCandidato) -> bool:
    try:
        return bool(
            await pai.elemento.evaluate(
                "(el, other) => el !== other && el.contains(other)",
                filho.elemento,
            )
        )
    except Exception:
        return False


async def remover_ancestrais_movimento(candidatos: list[MovimentoCandidato]) -> list[MovimentoCandidato]:
    if len(candidatos) < 2:
        return candidatos

    removidos: set[int] = set()

    for indice_pai, pai in enumerate(candidatos):
        for indice_filho, filho in enumerate(candidatos):
            if indice_pai == indice_filho:
                continue

            if not await elemento_contem_outro(pai, filho):
                continue

            if area_movimento(pai) >= area_movimento(filho) * 1.05:
                removidos.add(indice_pai)
                break

    return [candidato for indice, candidato in enumerate(candidatos) if indice not in removidos]


def remover_duplicatas_sobrepostas(candidatos: list[MovimentoCandidato]) -> list[MovimentoCandidato]:
    escolhidos: list[MovimentoCandidato] = []
    assinaturas: list[str] = []

    for candidato in sorted(candidatos, key=lambda item: (item.y, item.x, item.indice_original)):
        assinatura = assinatura_movimento(candidato.texto)
        substituido = False

        for indice, existente in enumerate(escolhidos):
            if assinatura != assinaturas[indice]:
                continue
            if not movimentos_sobrepostos(candidato, existente):
                continue

            if score_preferencia_movimento(candidato) < score_preferencia_movimento(existente):
                escolhidos[indice] = candidato
                assinaturas[indice] = assinatura
            substituido = True
            break

        if not substituido:
            escolhidos.append(candidato)
            assinaturas.append(assinatura)

    return sorted(escolhidos, key=lambda item: (item.y, item.x, item.indice_original))


async def filtrar_candidatos_movimento(candidatos: list[MovimentoCandidato]) -> list[MovimentoCandidato]:
    candidatos = await remover_ancestrais_movimento(candidatos)
    candidatos = remover_duplicatas_sobrepostas(candidatos)
    return candidatos


def categoria_nao_exige_venda(categoria: str) -> bool:
    return categoria.strip().lower() in CATEGORIAS_SEM_DETALHE


def chave_item_processado(pagina_atual: int, texto_item: str) -> str:
    digest = hashlib.sha1(texto_item.encode("utf-8", errors="ignore")).hexdigest()[:16]
    return f"pagina_{pagina_atual}_{digest}"


async def obter_data_formatada_do_item(
    pagina: Page,
    item: ElementoMovimento,
    texto_item: str,
    periodo: Periodo,
    data_fallback: str = "",
) -> str:
    data_texto = await obter_data_mais_proxima(pagina, item)

    if not data_texto:
        data_texto = extrair_data_extenso_do_texto(texto_item)

    if data_texto:
        return converter_data_extenso(data_texto, ano_base=periodo.inicio.year)

    if data_fallback:
        return data_fallback

    if periodo.inicio.date() == periodo.fim.date():
        return periodo.inicio.strftime("%d/%m/%Y")

    return ""


async def processar_pagina_atual(
    navegador: BrowserContext,
    pagina: Page,
    periodo: Periodo,
    pagina_atual: int,
    itens_processados: set[str],
) -> tuple[list[Venda], list[ErroItem]]:
    vendas: list[Venda] = []
    erros: list[ErroItem] = []
    lote_para_detalhar: list[tuple[ItemExtrato, ElementoMovimento]] = []

    linhas_movimento = await obter_linhas_movimento(pagina)
    total = len(linhas_movimento)

    log_info(f"Itens candidatos encontrados nesta página: {total}")

    if total == 0:
        log_warn("Nenhum movimento foi encontrado na página atual.")
        return vendas, erros

    ultima_data_formatada = (
        periodo.inicio.strftime("%d/%m/%Y")
        if periodo.inicio.date() == periodo.fim.date()
        else ""
    )

    for indice, movimento in enumerate(linhas_movimento):
        item = movimento.elemento
        texto_item = movimento.texto

        try:
            if item_deve_ser_ignorado(texto_item, descartar_container_data=False):
                log_debug(f"Item {indice + 1} ignorado por filtro de container/sem valor.")
                continue

            linhas_item = normalizar_linhas(texto_item)
            categoria_previa = identificar_categoria(linhas_item)

            if movimento_operacional_ignorado(texto_item, categoria_previa):
                log_debug(f"Item {indice + 1} ignorado por regra operacional: {categoria_previa}")
                continue

            chave_item = chave_item_processado(pagina_atual, texto_item)
            if chave_item in itens_processados:
                continue

            data_formatada = await obter_data_formatada_do_item(
                pagina,
                item,
                texto_item,
                periodo,
                data_fallback=ultima_data_formatada,
            )
            if data_formatada:
                ultima_data_formatada = data_formatada

            item_extraido = await extrair_dados_item(item, data_formatada, texto_item=texto_item)
            itens_processados.add(chave_item)

            console.print(
                f"[cyan]Item {indice + 1}/{total}[/cyan] | "
                f"{item_extraido.data_venda} | "
                f"{item_extraido.categoria} | "
                f"[bold]R$ {item_extraido.valor}[/bold]"
            )

            if categoria_nao_exige_venda(item_extraido.categoria):
                vendas.append(
                    Venda(
                        data_venda=item_extraido.data_venda,
                        numero_venda="",
                        categoria=item_extraido.categoria,
                        valor_liquido=item_extraido.valor,
                        valor_bruto="",
                    )
                )
                continue

            lote_para_detalhar.append((item_extraido, item))

            if len(lote_para_detalhar) >= MAX_ABAS_PARALELAS:
                vendas_lote, erros_lote = await processar_lote_itens(
                    navegador=navegador,
                    pagina=pagina,
                    lote=lote_para_detalhar,
                )
                vendas.extend(vendas_lote)
                erros.extend(erros_lote)
                lote_para_detalhar.clear()

        except Exception as erro:
            log_erro(f"Erro ao processar item {indice + 1}: {erro}")
            erros.append(
                ErroItem(
                    data_venda="",
                    categoria="",
                    valor="",
                    hora="",
                    motivo=f"Erro ao processar item {indice + 1} da página {pagina_atual}: {erro}",
                )
            )

    if lote_para_detalhar:
        vendas_lote, erros_lote = await processar_lote_itens(
            navegador=navegador,
            pagina=pagina,
            lote=lote_para_detalhar,
        )
        vendas.extend(vendas_lote)
        erros.extend(erros_lote)

    return vendas, erros


async def tentar_avancar_pagina(pagina: Page) -> bool:
    botao_proximo = pagina.get_by_text("Próximo", exact=True).first

    try:
        await botao_proximo.wait_for(state="visible", timeout=3_000)
    except PlaywrightTimeoutError:
        log_warn("Botão Próximo não encontrado. Fim da paginação.")
        return False

    try:
        classe = await botao_proximo.get_attribute("class") or ""
        aria_disabled = await botao_proximo.get_attribute("aria-disabled") or ""

        if "disabled" in classe.lower() or aria_disabled.lower() == "true":
            log_warn("Botão Próximo está desabilitado. Fim da paginação.")
            return False

        texto_antes = await pagina.locator("body").inner_text(timeout=TIMEOUT_CURTO)
        log_info("Clicando em Próximo...")
        await botao_proximo.click()

        try:
            await pagina.wait_for_function(
                "textoAntes => document.body && document.body.innerText !== textoAntes",
                arg=texto_antes,
                timeout=10_000,
            )
        except PlaywrightTimeoutError:
            log_debug("Texto da página não mudou rapidamente após Próximo; validando lista mesmo assim.")

        await aguardar_lista_movimentos(pagina)
        return True

    except Exception as erro:
        log_warn(f"Não consegui avançar para a próxima página: {erro}")
        return False


async def processar_todas_paginas(navegador: BrowserContext, pagina: Page, periodo: Periodo) -> ResultadoProcessamento:
    pagina_atual = 1
    itens_processados: set[str] = set()
    vendas: list[Venda] = []
    erros: list[ErroItem] = []

    while True:
        if MAX_PAGINAS and pagina_atual > MAX_PAGINAS:
            log_warn(f"Limite MAX_PAGINAS={MAX_PAGINAS} atingido. Encerrando paginação.")
            break

        console.print(Panel.fit(f"[bold]PROCESSANDO PÁGINA {pagina_atual}[/bold]", border_style="blue"))

        vendas_pagina, erros_pagina = await processar_pagina_atual(
            navegador=navegador,
            pagina=pagina,
            periodo=periodo,
            pagina_atual=pagina_atual,
            itens_processados=itens_processados,
        )

        vendas.extend(vendas_pagina)
        erros.extend(erros_pagina)
        log_ok(f"Página {pagina_atual} finalizada.")

        if not await tentar_avancar_pagina(pagina):
            break

        pagina_atual += 1

    return ResultadoProcessamento(vendas=vendas, erros=erros, total_itens=len(itens_processados))


# =============================================================================
# ERROS E RELATÓRIOS
# =============================================================================

def salvar_log_erros(erros: list[ErroItem], arquivos: ArquivosExecucao) -> None:
    if not erros:
        return

    arquivos.log_erros.parent.mkdir(parents=True, exist_ok=True)

    with open(arquivos.log_erros, "a", encoding="utf-8") as log:
        log.write("\n")
        log.write("=" * 80 + "\n")
        log.write(f"EXECUÇÃO: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
        log.write(f"ARQUIVO EXCEL: {arquivos.planilha}\n")
        log.write("=" * 80 + "\n\n")

        for erro in erros:
            log.write(
                f"Data: {erro.data_venda} | "
                f"Hora: {erro.hora} | "
                f"Categoria: {erro.categoria} | "
                f"Valor: {erro.valor} | "
                f"Motivo: {erro.motivo}\n"
            )


def exibir_erros(erros: list[ErroItem]) -> None:
    if not erros:
        return

    tabela = Table(title="Itens com erro / não exportados")
    tabela.add_column("Data", style="cyan")
    tabela.add_column("Hora", style="cyan")
    tabela.add_column("Categoria")
    tabela.add_column("Valor", justify="right")
    tabela.add_column("Motivo", style="red")

    for erro in erros:
        tabela.add_row(erro.data_venda, erro.hora, erro.categoria, erro.valor, erro.motivo)

    console.print(tabela)


def exibir_resumo_final(
    total_itens: int,
    total_vendas: int,
    total_erros: int,
    total_exportado: int,
    inicio_execucao: float,
) -> None:
    tempo_total = int(time.perf_counter() - inicio_execucao)
    minutos = tempo_total // 60
    segundos = tempo_total % 60

    console.print(
        Panel.fit(
            f"[bold green]Processo finalizado[/bold green]\n"
            f"Itens processados: [bold]{total_itens}[/bold]\n"
            f"Vendas capturadas: [bold]{total_vendas}[/bold]\n"
            f"Vendas exportadas: [bold]{total_exportado}[/bold]\n"
            f"Erros: [bold]{total_erros}[/bold]\n"
            f"Tempo total: [bold]{minutos} minuto(s) e {segundos} segundo(s)[/bold]",
            border_style="green",
        )
    )


# =============================================================================
# ORQUESTRAÇÃO
# =============================================================================

def preparar_terminal() -> None:
    if sys.platform.startswith("win"):
        os.system("mode con cols=120 lines=40")
        os.system("cls")


async def executar_robo() -> None:
    preparar_terminal()

    inicio_execucao = time.perf_counter()
    periodo = calcular_periodo()
    arquivos = definir_arquivos_execucao(periodo)

    console.print(Panel.fit(f"[bold]Arquivo de saída[/bold]\n{arquivos.planilha}", border_style="cyan"))

    navegador: BrowserContext | None = None

    async with async_playwright() as playwright:
        try:
            navegador = await abrir_navegador(playwright)
            pagina = await navegador.new_page()
            pagina.set_default_timeout(TIMEOUT_MEDIO)

            await fechar_abas_em_branco(navegador, pagina)
            await abrir_mercado_pago(pagina)
            await selecionar_periodo(pagina, periodo)
            await filtrar_todas_entradas(pagina)
            await aguardar_lista_movimentos(pagina)

            resultado = await processar_todas_paginas(navegador, pagina, periodo)

            total_exportado = exportar_vendas_para_planilha(resultado.vendas, arquivos.planilha)
            exibir_erros(resultado.erros)
            salvar_log_erros(resultado.erros, arquivos)

            if resultado.erros:
                log_warn(f"Log de erros salvo em: {arquivos.log_erros}")

            exibir_resumo_final(
                total_itens=resultado.total_itens,
                total_vendas=len(resultado.vendas),
                total_erros=len(resultado.erros),
                total_exportado=total_exportado,
                inicio_execucao=inicio_execucao,
            )
            console.print("[green]Processo finalizado. Fechando automaticamente em 5 segundos...[/green]")
            await asyncio.sleep(5)

        except PlaywrightTimeoutError as erro:
            log_erro("Erro de timeout:")
            console.print(erro)
            input("O navegador ficará aberto até você pressionar ENTER...")

        except Exception as erro:
            log_erro("Deu erro:")
            console.print(erro)
            input("O navegador ficará aberto até você pressionar ENTER...")

        finally:
            if navegador is not None:
                await navegador.close()


def main() -> None:
    asyncio.run(executar_robo())


if __name__ == "__main__":
    main()
